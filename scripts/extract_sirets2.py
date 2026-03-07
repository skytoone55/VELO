import openpyxl, re, json
from datetime import datetime, date

wb = openpyxl.load_workbook("/Users/john/Library/CloudStorage/Dropbox/VELO CARGO/ENEMAT/PPE - TRA-EQ-131 cloture.xlsx", data_only=True)
ws = wb.active

# Check all sheets in workbook
print("All sheets:", wb.sheetnames)
print(f"Active sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")

# Sample first 5 data rows to see what's in each column
for row in range(2, min(7, ws.max_row+1)):
    print(f"\n--- Row {row} ---")
    col_t_val = ws.cell(row=row, column=20).value  # T = col 20
    print(f"  Col T (professionnel): {col_t_val}")
    col_g_val = ws.cell(row=row, column=7).value   # G = col 7
    print(f"  Col G (date engagement): {col_g_val}")
    col_e_val = ws.cell(row=row, column=5).value   # E = col 5
    print(f"  Col E (date envoi RAI): {col_e_val}")
    col_c_val = ws.cell(row=row, column=3).value   # C = col 3
    print(f"  Col C (SIREN demandeur): {col_c_val}")
    col_m_val = ws.cell(row=row, column=13).value  # M = col 13
    print(f"  Col M (SIREN beneficiaire): {col_m_val}")
    col_l_val = ws.cell(row=row, column=12).value  # L = col 12
    print(f"  Col L (raison sociale benef): {col_l_val}")
    col_p_val = ws.cell(row=row, column=16).value  # P = col 16
    print(f"  Col P (ville): {col_p_val}")
    col_o_val = ws.cell(row=row, column=15).value  # O = col 15
    print(f"  Col O (CP): {col_o_val}")

# Also check if there are more sheets with a SIRET column
for sn in wb.sheetnames:
    sheet = wb[sn]
    print(f"\nSheet '{sn}': {sheet.max_row} rows x {sheet.max_column} cols")
    if sheet.max_column >= 26:
        print(f"  Col Z value row1: {sheet.cell(row=1, column=26).value}")
        print(f"  Col Z value row2: {sheet.cell(row=1, column=26).value}")

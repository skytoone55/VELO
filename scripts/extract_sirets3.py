import openpyxl, re

wb = openpyxl.load_workbook("/Users/john/Library/CloudStorage/Dropbox/VELO CARGO/ENEMAT/PPE - TRA-EQ-131 cloture.xlsx", data_only=True)

for sn in ['01-10-2025', '19-09-2025']:
    sheet = wb[sn]
    print(f"\n=== Sheet '{sn}': {sheet.max_row} rows x {sheet.max_column} cols ===")
    for c in range(1, min(sheet.max_column+1, 28)):
        v = sheet.cell(row=1, column=c).value
        letter = openpyxl.utils.get_column_letter(c)
        print(f"  Col {c} ({letter}): {v}")
    # Sample row 2
    print(f"\n  Sample row 2:")
    print(f"    Col T (20): {sheet.cell(row=2, column=20).value}")
    print(f"    Col G (7):  {sheet.cell(row=2, column=7).value}")
    print(f"    Col E (5):  {sheet.cell(row=2, column=5).value}")
    print(f"    Col Z (26): {sheet.cell(row=2, column=26).value}")
    print(f"    Col L (12): {sheet.cell(row=2, column=12).value}")
    print(f"    Col P (16): {sheet.cell(row=2, column=16).value}")
    print(f"    Col O (15): {sheet.cell(row=2, column=15).value}")

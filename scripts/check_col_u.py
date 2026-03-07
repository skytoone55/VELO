import openpyxl, json, re
from collections import Counter

# Load missing SIRETs from ENEMAT
with open('/Users/john/JARVIS/enemat_ppe_sirets.json') as f:
    enemat_data = json.load(f)

# Load current Monday SIRETs
with open('/Users/john/JARVIS/monday_sirets.json') as f:
    monday_data = json.load(f)

print(f"ENEMAT SIRETs loaded: {len(enemat_data)}")
print(f"Monday SIRETs loaded: {len(monday_data)}")

# Read old Excel file
wb = openpyxl.load_workbook("/Users/john/JARVIS/velo/donnee excel.xlsx", read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows_all = list(ws.iter_rows(min_row=1, values_only=True))
headers = rows_all[0]

print(f"Total rows (incl header): {len(rows_all)}")
print(f"Number of columns: {len(headers)}")
print()

# Print column U and nearby columns for context
for i in range(18, 24):
    if i < len(headers):
        col_letter = chr(65+i) if i < 26 else 'A' + chr(65+i-26)
        print(f"  Col {col_letter} (index {i}): {headers[i]}")

print()

# Column U = index 20
# Column Z = index 25 (SIRET)
# Read ALL data rows, build a dict by cleaned SIRET
excel_by_siret = {}
for r in rows_all[1:]:
    siret_raw = str(r[25] or '') if len(r) > 25 else ''
    siret = re.sub(r'\D', '', siret_raw)
    if len(siret) >= 9:
        siret = siret.zfill(14) if len(siret) < 14 else siret
        col_u = str(r[20] or '').strip() if len(r) > 20 else ''
        excel_by_siret[siret] = col_u

print(f"Excel rows with valid SIRET: {len(excel_by_siret)}")

# Compute missing SIRETs (ENEMAT - Monday)
monday_sirets_clean = set()
for s in monday_data.keys():
    monday_sirets_clean.add(re.sub(r'\D', '', s).zfill(14))

enemat_sirets_clean = set()
for s in enemat_data.keys():
    enemat_sirets_clean.add(re.sub(r'\D', '', s).zfill(14))

missing_sirets = enemat_sirets_clean - monday_sirets_clean
print(f"Missing SIRETs (ENEMAT - Monday): {len(missing_sirets)}")

# For each missing SIRET, get column U value from old Excel
col_u_values = Counter()
found_in_excel = 0
not_found = 0
not_found_list = []
for siret in sorted(missing_sirets):
    if siret in excel_by_siret:
        found_in_excel += 1
        val = excel_by_siret[siret] if excel_by_siret[siret] else '(vide)'
        col_u_values[val] += 1
    else:
        not_found += 1
        if len(not_found_list) < 10:
            not_found_list.append(siret)

print(f"Found in old Excel: {found_in_excel}")
print(f"Not found in old Excel: {not_found}")
if not_found_list:
    print(f"  (first 10 not found: {not_found_list})")

print(f"\n=== Colonne U — Repartition des {found_in_excel} clients manquants ===\n")

for val, count in col_u_values.most_common():
    pct = count / found_in_excel * 100 if found_in_excel > 0 else 0
    print(f"  {val}: {count} ({pct:.1f}%)")

wb.close()
print("\nDone. READ-ONLY, no modifications made.")

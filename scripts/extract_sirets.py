import openpyxl, re, json
from datetime import datetime, date

wb = openpyxl.load_workbook("/Users/john/Library/CloudStorage/Dropbox/VELO CARGO/ENEMAT/PPE - TRA-EQ-131 cloture.xlsx", data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")

# Print headers to understand structure
for c in range(1, min(ws.max_column+1, 30)):
    v = ws.cell(row=1, column=c).value
    letter = openpyxl.utils.get_column_letter(c)
    print(f"  Col {c-1} ({letter}): {v}")

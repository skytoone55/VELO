#!/usr/bin/env python3
"""
ECOVOLT - Stats Validation NAF ENEMAT
Extrait les donnees du board Monday.com Velos Cargos General
et genere un fichier Excel professionnel.
"""

import json
import time
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_URL = "https://api.monday.com/v2"
API_TOKEN = "REDACTED_MONDAY_ECOVOLT_TOKEN"
BOARD_ID = 9990833105
COLUMNS = [
    "color_mkvfws5n",
    "numeric_mkvfghjq",
    "numeric_mkvcqm0r",
    "color_mkvdkzxh",
    "text_mkvft2w3",
    "color_mm0vxv46",
]
COL_IDS_STR = ", ".join(f'"{c}"' for c in COLUMNS)

NAF_REF_PATH = Path("/Users/john/JARVIS/projets/velo/docs/naf-enemat-reference.json")
OUTPUT_PATH = Path("/Users/john/JARVIS/CLAUDE/Stats_Ecovolt.xlsx")

VAL_NAF_MAP = {
    "Fait": "OUI",
    "Bloqué": "NON",
    "En cours": "A VERIFIER",
}

def map_ile(dept_text):
    if not dept_text:
        return "Autre"
    d = dept_text.lower()
    if "martinique" in d:
        return "Martinique"
    if "guadeloupe" in d:
        return "Guadeloupe"
    if "guyane" in d:
        return "Guyane"
    if "reunion" in d or "réunion" in d:
        return "Réunion"
    if "mayotte" in d:
        return "Mayotte"
    return "Autre"

ILE_ORDER = ["Martinique", "Guadeloupe", "Guyane", "Réunion", "Mayotte", "Autre"]


def monday_query(query):
    payload = json.dumps({"query": query}).encode("utf-8")
    req = urllib.request.Request(
        API_URL,
        data=payload,
        headers={
            "Authorization": API_TOKEN,
            "Content-Type": "application/json",
            "API-Version": "2024-10",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_all_items():
    items = []
    q1 = '{ boards(ids: [' + str(BOARD_ID) + ']) { items_page(limit: 500) { cursor items { id name column_values(ids: [' + COL_IDS_STR + ']) { id text value } } } } }'
    print("Requete page 1...")
    data = monday_query(q1)
    if "errors" in data:
        print("ERREUR API:", data["errors"])
        raise SystemExit(1)

    page = data["data"]["boards"][0]["items_page"]
    items.extend(page["items"])
    cursor = page.get("cursor")
    print(f"  -> {len(items)} items recuperes")

    page_num = 1
    while cursor:
        page_num += 1
        time.sleep(0.5)
        qn = '{ next_items_page(limit: 500, cursor: "' + cursor + '") { cursor items { id name column_values(ids: [' + COL_IDS_STR + ']) { id text value } } } }'
        print(f"Requete page {page_num}...")
        data = monday_query(qn)
        if "errors" in data:
            print("ERREUR API:", data["errors"])
            raise SystemExit(1)
        page = data["data"]["next_items_page"]
        items.extend(page["items"])
        cursor = page.get("cursor")
        print(f"  -> {len(items)} items recuperes au total")

    return items


def parse_items(raw_items):
    records = []
    for item in raw_items:
        cols = {cv["id"]: cv.get("text", "") or "" for cv in item["column_values"]}

        statut = cols.get("color_mkvfws5n", "").strip()
        naf_code = cols.get("text_mkvft2w3", "").strip()
        val_naf_raw = cols.get("color_mm0vxv46", "").strip()
        dept_raw = cols.get("color_mkvdkzxh", "").strip()

        def safe_int(v):
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return 0

        velo_confirme = safe_int(cols.get("numeric_mkvcqm0r", ""))
        velo_devis = safe_int(cols.get("numeric_mkvfghjq", ""))
        velos = velo_confirme if velo_confirme > 0 else velo_devis

        validation_naf = VAL_NAF_MAP.get(val_naf_raw, "A VERIFIER")
        ile = map_ile(dept_raw)
        is_cv = (statut == "CONTROLE VALIDÉ")

        records.append({
            "id": item["id"],
            "name": item["name"],
            "statut": statut,
            "naf_code": naf_code,
            "validation_naf": validation_naf,
            "velos": velos,
            "departement": dept_raw,
            "ile": ile,
            "is_cv": is_cv,
            "is_ok": validation_naf == "OUI",
            "is_ko": validation_naf == "NON",
            "has_1_velo": velos == 1,
            "has_multi_velo": velos > 1,
        })
    return records


def compute_stats(records):
    total_clients = len(records)
    total_velos = sum(r["velos"] for r in records)

    clients_ok = sum(1 for r in records if r["is_ok"])
    clients_ko = sum(1 for r in records if r["is_ko"])
    clients_averif = total_clients - clients_ok - clients_ko

    velos_ok = sum(r["velos"] for r in records if r["is_ok"])
    velos_ko = sum(r["velos"] for r in records if r["is_ko"])
    velos_averif = total_velos - velos_ok - velos_ko

    c1_ok = sum(1 for r in records if r["has_1_velo"] and r["is_ok"])
    c1_ko = sum(1 for r in records if r["has_1_velo"] and r["is_ko"])
    v1_ok = sum(r["velos"] for r in records if r["has_1_velo"] and r["is_ok"])
    v1_ko = sum(r["velos"] for r in records if r["has_1_velo"] and r["is_ko"])

    cm_ok = sum(1 for r in records if r["has_multi_velo"] and r["is_ok"])
    cm_ko = sum(1 for r in records if r["has_multi_velo"] and r["is_ko"])
    vm_ok = sum(r["velos"] for r in records if r["has_multi_velo"] and r["is_ok"])
    vm_ko = sum(r["velos"] for r in records if r["has_multi_velo"] and r["is_ko"])

    naf_ko_data = defaultdict(lambda: {"clients": 0, "velos": 0, "c1": 0, "cm": 0, "v1": 0, "vm": 0})
    for r in records:
        if r["is_ko"]:
            code = r["naf_code"] if r["naf_code"] else "(vide)"
            naf_ko_data[code]["clients"] += 1
            naf_ko_data[code]["velos"] += r["velos"]
            if r["has_1_velo"]:
                naf_ko_data[code]["c1"] += 1
                naf_ko_data[code]["v1"] += r["velos"]
            if r["has_multi_velo"]:
                naf_ko_data[code]["cm"] += 1
                naf_ko_data[code]["vm"] += r["velos"]

    naf_ko_list = sorted(naf_ko_data.items(), key=lambda x: x[1]["velos"], reverse=True)

    geo = {}
    for ile in ILE_ORDER:
        sub = [r for r in records if r["ile"] == ile]
        geo[ile] = {
            "clients_total": len(sub),
            "clients_ok": sum(1 for r in sub if r["is_ok"]),
            "clients_ko": sum(1 for r in sub if r["is_ko"]),
            "velos_total": sum(r["velos"] for r in sub),
            "velos_ok": sum(r["velos"] for r in sub if r["is_ok"]),
            "velos_ko": sum(r["velos"] for r in sub if r["is_ko"]),
        }

    return {
        "total_clients": total_clients,
        "total_velos": total_velos,
        "clients_ok": clients_ok,
        "clients_ko": clients_ko,
        "clients_averif": clients_averif,
        "velos_ok": velos_ok,
        "velos_ko": velos_ko,
        "velos_averif": velos_averif,
        "c1_ok": c1_ok, "c1_ko": c1_ko,
        "cm_ok": cm_ok, "cm_ko": cm_ko,
        "v1_ok": v1_ok, "v1_ko": v1_ko,
        "vm_ok": vm_ok, "vm_ko": vm_ko,
        "naf_ko": naf_ko_list,
        "geo": geo,
    }


# ===== EXCEL STYLES =====
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="333333")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="666666")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="E07020")

FILL_HEADER = PatternFill(start_color="E07020", end_color="E07020", fill_type="solid")
FILL_ALT = PatternFill(start_color="FFF5EB", end_color="FFF5EB", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FFE0C0", end_color="FFE0C0", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        length = min_w
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[col_letter].width = length


def style_row(ws, row, num_cols, font=FONT_NORMAL, fill=FILL_WHITE, align=ALIGN_CENTER):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = align


def pct(part, total):
    if total == 0:
        return chr(8211)  # en-dash
    return f"{part/total*100:.1f}%"


def write_synthese(ws, stats_cv, stats_other):
    ws.merge_cells("A1:H1")
    ws["A1"].value = "ECOVOLT \u2014 Statistiques Validation NAF ENEMAT"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Date : 23/02/2026 | Board : V\u00e9los Cargos G\u00e9n\u00e9ral | P\u00e9rim\u00e8tre DOM-TOM"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_LEFT

    def write_section(ws, start_row, title, s):
        r = start_row
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value=title).font = FONT_SECTION
        r += 1

        headers = ["Indicateur", "Total", "OK (\u00e9ligible)", "% OK", "KO (non \u00e9ligible)", "% KO", "\u00c0 v\u00e9rifier", "% \u00c0 v\u00e9rif."]
        for i, h in enumerate(headers, 1):
            ws.cell(row=r, column=i, value=h)
        style_row(ws, r, 8, FONT_HEADER, FILL_HEADER)
        r += 1

        rows_data = [
            ("Clients \u2014 Total",
             s["total_clients"], s["clients_ok"], pct(s["clients_ok"], s["total_clients"]),
             s["clients_ko"], pct(s["clients_ko"], s["total_clients"]),
             s["clients_averif"], pct(s["clients_averif"], s["total_clients"])),

            ("V\u00e9los \u2014 Total",
             s["total_velos"], s["velos_ok"], pct(s["velos_ok"], s["total_velos"]),
             s["velos_ko"], pct(s["velos_ko"], s["total_velos"]),
             s["velos_averif"], pct(s["velos_averif"], s["total_velos"])),

            ("Clients \u00e0 1 v\u00e9lo",
             s["c1_ok"] + s["c1_ko"], s["c1_ok"], pct(s["c1_ok"], s["c1_ok"] + s["c1_ko"]),
             s["c1_ko"], pct(s["c1_ko"], s["c1_ok"] + s["c1_ko"]),
             "", ""),

            ("Clients \u00e0 +1 v\u00e9los",
             s["cm_ok"] + s["cm_ko"], s["cm_ok"], pct(s["cm_ok"], s["cm_ok"] + s["cm_ko"]),
             s["cm_ko"], pct(s["cm_ko"], s["cm_ok"] + s["cm_ko"]),
             "", ""),

            ("V\u00e9los (clients 1 v\u00e9lo)",
             s["v1_ok"] + s["v1_ko"], s["v1_ok"], pct(s["v1_ok"], s["v1_ok"] + s["v1_ko"]),
             s["v1_ko"], pct(s["v1_ko"], s["v1_ok"] + s["v1_ko"]),
             "", ""),

            ("V\u00e9los (clients +1 v\u00e9los)",
             s["vm_ok"] + s["vm_ko"], s["vm_ok"], pct(s["vm_ok"], s["vm_ok"] + s["vm_ko"]),
             s["vm_ko"], pct(s["vm_ko"], s["vm_ok"] + s["vm_ko"]),
             "", ""),
        ]

        for idx, row_data in enumerate(rows_data):
            for i, val in enumerate(row_data, 1):
                ws.cell(row=r, column=i, value=val)
            fill = FILL_ALT if idx % 2 == 1 else FILL_WHITE
            style_row(ws, r, 8, FONT_NORMAL, fill)
            ws.cell(row=r, column=1).alignment = ALIGN_LEFT
            r += 1

        return r + 1

    next_row = write_section(ws, 4, "\u25b8 Clients CONTROLE VALID\u00c9", stats_cv)
    write_section(ws, next_row, "\u25b8 Clients hors CONTROLE VALID\u00c9 (en attente)", stats_other)

    ws.freeze_panes = "A4"
    auto_width(ws)


def write_naf_ko(ws, title, naf_ko_list, naf_labels):
    ws.merge_cells("A1:H1")
    ws["A1"].value = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT

    r = 3
    headers = ["Code NAF", "Libell\u00e9 activit\u00e9", "Clients KO", "V\u00e9los KO",
               "Clients 1 v\u00e9lo", "Clients +1 v\u00e9lo", "V\u00e9los (1 v\u00e9lo)", "V\u00e9los (+1 v\u00e9los)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_row(ws, r, 8, FONT_HEADER, FILL_HEADER)
    r += 1

    total_clients = 0
    total_velos = 0
    total_c1 = 0
    total_cm = 0
    total_v1 = 0
    total_vm = 0

    for idx, (code, d) in enumerate(naf_ko_list):
        label = naf_labels.get(code, "\u2014")
        row_data = [code, label, d["clients"], d["velos"], d["c1"], d["cm"], d["v1"], d["vm"]]
        for i, val in enumerate(row_data, 1):
            ws.cell(row=r, column=i, value=val)
        fill = FILL_ALT if idx % 2 == 1 else FILL_WHITE
        style_row(ws, r, 8, FONT_NORMAL, fill)
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT

        total_clients += d["clients"]
        total_velos += d["velos"]
        total_c1 += d["c1"]
        total_cm += d["cm"]
        total_v1 += d["v1"]
        total_vm += d["vm"]
        r += 1

    total_row = ["TOTAL", "", total_clients, total_velos, total_c1, total_cm, total_v1, total_vm]
    for i, val in enumerate(total_row, 1):
        ws.cell(row=r, column=i, value=val)
    style_row(ws, r, 8, FONT_BOLD, FILL_TOTAL)
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT

    ws.freeze_panes = "A4"
    auto_width(ws, max_w=55)


def write_geo(ws, title, geo_data):
    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT

    r = 3
    headers = ["\u00cele", "Clients Total", "Clients OK", "Clients KO",
               "V\u00e9los Total", "V\u00e9los OK", "V\u00e9los KO"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_row(ws, r, 7, FONT_HEADER, FILL_HEADER)
    r += 1

    totals = {"clients_total": 0, "clients_ok": 0, "clients_ko": 0,
              "velos_total": 0, "velos_ok": 0, "velos_ko": 0}

    idx = 0
    for ile in ILE_ORDER:
        if ile not in geo_data:
            continue
        d = geo_data[ile]
        row_data = [ile, d["clients_total"], d["clients_ok"], d["clients_ko"],
                    d["velos_total"], d["velos_ok"], d["velos_ko"]]
        for i, val in enumerate(row_data, 1):
            ws.cell(row=r, column=i, value=val)
        fill = FILL_ALT if idx % 2 == 1 else FILL_WHITE
        style_row(ws, r, 7, FONT_NORMAL, fill)
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT

        for k in totals:
            totals[k] += d[k]
        idx += 1
        r += 1

    total_row = ["TOTAL", totals["clients_total"], totals["clients_ok"], totals["clients_ko"],
                 totals["velos_total"], totals["velos_ok"], totals["velos_ko"]]
    for i, val in enumerate(total_row, 1):
        ws.cell(row=r, column=i, value=val)
    style_row(ws, r, 7, FONT_BOLD, FILL_TOTAL)
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT

    ws.freeze_panes = "A4"
    auto_width(ws)


def main():
    print("=" * 60)
    print("ECOVOLT - Extraction & Stats Validation NAF ENEMAT")
    print("=" * 60)

    print("\\nChargement des labels NAF...")
    with open(NAF_REF_PATH, "r", encoding="utf-8") as f:
        naf_ref = json.load(f)
    naf_labels = naf_ref.get("labels", {})
    print(f"  -> {len(naf_labels)} codes NAF charges")

    print("\\nExtraction des items du board Ecovolt...")
    raw_items = fetch_all_items()
    print(f"\\nTotal items extraits : {len(raw_items)}")

    print("\\nTraitement des donnees...")
    records = parse_items(raw_items)

    rec_cv = [r for r in records if r["is_cv"]]
    rec_other = [r for r in records if not r["is_cv"]]
    print(f"  -> CONTROLE VALIDE : {len(rec_cv)} clients")
    print(f"  -> Hors CONTROLE VALIDE : {len(rec_other)} clients")

    stats_cv = compute_stats(rec_cv)
    stats_other = compute_stats(rec_other)

    print("\\n" + "=" * 60)
    print("RESUME - CONTROLE VALIDE")
    print(f"  Clients : {stats_cv['total_clients']} (OK: {stats_cv['clients_ok']}, KO: {stats_cv['clients_ko']}, A verifier: {stats_cv['clients_averif']})")
    print(f"  Velos   : {stats_cv['total_velos']} (OK: {stats_cv['velos_ok']}, KO: {stats_cv['velos_ko']}, A verifier: {stats_cv['velos_averif']})")
    print(f"  NAF KO distincts : {len(stats_cv['naf_ko'])}")

    print("\\nRESUME - HORS CONTROLE VALIDE")
    print(f"  Clients : {stats_other['total_clients']} (OK: {stats_other['clients_ok']}, KO: {stats_other['clients_ko']}, A verifier: {stats_other['clients_averif']})")
    print(f"  Velos   : {stats_other['total_velos']} (OK: {stats_other['velos_ok']}, KO: {stats_other['velos_ko']}, A verifier: {stats_other['velos_averif']})")
    print(f"  NAF KO distincts : {len(stats_other['naf_ko'])}")

    print("\\nVENTILATION PAR ILE - CONTROLE VALIDE")
    for ile in ILE_ORDER:
        g = stats_cv["geo"].get(ile, {})
        if g.get("clients_total", 0) > 0:
            print(f"  {ile:15s} : {g['clients_total']:4d} clients ({g['clients_ok']} OK, {g['clients_ko']} KO) | {g['velos_total']:4d} velos ({g['velos_ok']} OK, {g['velos_ko']} KO)")

    print("\\nVENTILATION PAR ILE - HORS CONTROLE VALIDE")
    for ile in ILE_ORDER:
        g = stats_other["geo"].get(ile, {})
        if g.get("clients_total", 0) > 0:
            print(f"  {ile:15s} : {g['clients_total']:4d} clients ({g['clients_ok']} OK, {g['clients_ko']} KO) | {g['velos_total']:4d} velos ({g['velos_ok']} OK, {g['velos_ko']} KO)")

    print("=" * 60)

    print("\\nGeneration du fichier Excel...")
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "SYNTHESE"
    write_synthese(ws1, stats_cv, stats_other)

    ws2 = wb.create_sheet("NAF KO - CONTROLE VALIDE")
    write_naf_ko(ws2, "Codes NAF Non \u00c9ligibles \u2014 Clients CONTROLE VALID\u00c9", stats_cv["naf_ko"], naf_labels)

    ws3 = wb.create_sheet("NAF KO - EN ATTENTE")
    write_naf_ko(ws3, "Codes NAF Non \u00c9ligibles \u2014 Clients hors CONTROLE VALID\u00c9", stats_other["naf_ko"], naf_labels)

    ws4 = wb.create_sheet("ILES - CONTROLE VALIDE")
    write_geo(ws4, "Ventilation par \u00eele \u2014 Clients CONTROLE VALID\u00c9", stats_cv["geo"])

    ws5 = wb.create_sheet("ILES - EN ATTENTE")
    write_geo(ws5, "Ventilation par \u00eele \u2014 Clients hors CONTROLE VALID\u00c9", stats_other["geo"])

    ws1.sheet_properties.tabColor = "E07020"
    ws2.sheet_properties.tabColor = "CC0000"
    ws3.sheet_properties.tabColor = "FFD700"
    ws4.sheet_properties.tabColor = "228B22"
    ws5.sheet_properties.tabColor = "808080"

    wb.save(str(OUTPUT_PATH))
    print(f"\\nFichier genere : {OUTPUT_PATH}")
    print("Termine.")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
V3: Corrected field mapping. Compare Ecovolt Excel vs Monday board.
Monday column mapping (corrected):
  name                     = beneficiary name (= Excel "RAISON SOCIALE du bénéficiaire")
  text_mkvfykn9            = SIRET_RETINA (beneficiary SIRET = Excel "SIREN du bénéficiaire" col 12)
  email_mkvfnv4q           = emailbeneficiaire_RETINA (= Excel "courriel du bénéficiaire" col 17)
  long_text_mkvn5k9w       = Telephonebeneficiaire_RETINA (= Excel "téléphone du bénéficiaire" col 16)
  text_mkvfetg2            = adresseopération_RETINA (= Excel "ADRESSE du siège social" col 13)
  text_mkvfhcn9            = CPoperation_RETINA (= Excel "CODE POSTAL" col 14)
  text_mkvfgh8t            = Villeopération_RETINA (= Excel "VILLE" col 15)
  text_mkvft2w3            = APE/NAF_RETINA (no direct Excel equivalent)
  text_mkvf8zp6            = Numerodevis_RETINA (= Excel "REFERENCE de la preuve" col 20)
  text_mkvfxbkp            = refinternedeloperation_RETINA (= Excel "REFERENCE interne" col 3)
"""
import json, sys, re, time
import openpyxl, requests

EXCEL_PATH = "/Users/john/JARVIS/projets/velo/docs/PPE - TRA-EQ-131 cloture-complet.xlsx"
BOARD_ID = 9990833105

with open("/Users/john/JARVIS/projets/velo/.env.local") as f:
    for line in f:
        if line.startswith("MONDAY_API_KEY"):
            API_KEY = line.split("=", 1)[1].strip().strip('"')
            break

MONDAY_URL = "https://api.monday.com/v2"
HEADERS = {"Authorization": API_KEY, "Content-Type": "application/json"}

MONDAY_COL_MAP = {
    "text_mkvfykn9": "siret_benef",
    "email_mkvfnv4q": "email_benef",
    "long_text_mkvn5k9w": "telephone",
    "text_mkvfetg2": "adresse",
    "text_mkvfhcn9": "code_postal",
    "text_mkvfgh8t": "ville",
    "text_mkvft2w3": "naf",
    "text_mkvf8zp6": "ref_devis",
    "text_mkvfxbkp": "ref_interne",
    "email_mkvfk63f": "email_agent",
}

def normalize(s):
    if not s: return ""
    s = str(s).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'\.0$', '', s)
    return s

def norm_siren(s):
    """Normalize SIREN/SIRET: remove spaces/dots, strip .0, compare first 9 digits."""
    if not s: return ""
    s = str(s).strip()
    s = re.sub(r'\.0$', '', s)
    s = re.sub(r'[^0-9]', '', s)
    return s

def norm_phone(s):
    if not s: return ""
    return re.sub(r'[^0-9+]', '', str(s))

def norm_cp(s):
    if not s: return ""
    s = str(s).strip()
    s = re.sub(r'\.0$', '', s)
    return s.zfill(5) if s.isdigit() and len(s) < 5 else s

def extract_ecovolt():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    entries = []
    sheet_counts = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheet_counts[sn] = 0
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower().replace("\n"," ").replace("\r"," ")
            if "raison sociale" in hl and "professionnel" in hl: col_map["rs_pro"] = i
            elif "raison sociale" in hl and "bénéficiaire" in hl: col_map["rs_benef"] = i
            elif "opération n" in hl: col_map["op_num"] = i
            elif "siren" in hl and "professionnel" in hl: col_map["siren_pro"] = i
            elif "siren" in hl and "bénéficiaire" in hl: col_map["siren_benef"] = i
            elif "reference interne" in hl: col_map["ref_interne"] = i
            elif "adresse" in hl and "siège" in hl: col_map["adresse"] = i
            elif "code postal" in hl: col_map["code_postal"] = i
            elif "ville" in hl: col_map["ville"] = i
            elif "téléphone" in hl or "telephone" in hl: col_map["telephone"] = i
            elif "courriel" in hl: col_map["email_benef"] = i
            elif "reference" in hl and "preuve" in hl: col_map["ref_preuve"] = i

        cnt = 0
        for row in rows[1:]:
            idx = col_map.get("rs_pro")
            if idx is None or idx >= len(row): continue
            rs = str(row[idx] or "").strip()
            if "eco-volt" not in rs.lower() and "ecovolt" not in rs.lower(): continue
            def gv(k):
                ii = col_map.get(k)
                if ii is not None and ii < len(row):
                    v = row[ii]
                    return str(v).strip() if v is not None else ""
                return ""
            entries.append({
                "sheet": sn, "op_num": gv("op_num"),
                "rs_benef": gv("rs_benef"),
                "siren_benef": gv("siren_benef"),
                "ref_interne": gv("ref_interne"),
                "adresse": gv("adresse"),
                "code_postal": gv("code_postal"),
                "ville": gv("ville"),
                "telephone": gv("telephone"),
                "email_benef": gv("email_benef"),
                "ref_preuve": gv("ref_preuve"),
            })
            cnt += 1
        sheet_counts[sn] = cnt
    wb.close()
    return entries, sheet_counts

def fetch_monday():
    items = []
    q = '{ boards(ids: %d) { items_page(limit: 500) { cursor items { id name column_values { id text value } } } } }' % BOARD_ID
    r = requests.post(MONDAY_URL, headers=HEADERS, json={"query": q}, timeout=30).json()
    pg = r["data"]["boards"][0]["items_page"]
    items.extend(pg["items"])
    cursor = pg["cursor"]
    while cursor:
        q2 = '{ next_items_page(limit: 500, cursor: "%s") { cursor items { id name column_values { id text value } } } }' % cursor
        r2 = requests.post(MONDAY_URL, headers=HEADERS, json={"query": q2}, timeout=30).json()
        pg2 = r2["data"]["next_items_page"]
        items.extend(pg2["items"])
        cursor = pg2["cursor"]
        time.sleep(0.3)
    result = []
    for it in items:
        d = {"id": it["id"], "name": it["name"]}
        for cv in it["column_values"]:
            if cv["id"] in MONDAY_COL_MAP:
                d[MONDAY_COL_MAP[cv["id"]]] = (cv.get("text") or "").strip()
        result.append(d)
    return result

def main():
    print("Excel...", file=sys.stderr)
    excel, sc = extract_ecovolt()
    print(f"  {len(excel)} entries {sc}", file=sys.stderr)
    print("Monday...", file=sys.stderr)
    monday = fetch_monday()
    print(f"  {len(monday)} items", file=sys.stderr)

    # Build Monday lookup by normalized name
    mon_by_name = {}
    for m in monday:
        k = normalize(m["name"])
        mon_by_name.setdefault(k, []).append(m)

    # Also build by ref_interne for fallback matching
    mon_by_ref = {}
    for m in monday:
        ref = normalize(m.get("ref_interne", ""))
        if ref:
            mon_by_ref.setdefault(ref, []).append(m)

    matched = []
    missing = []
    used_ids = set()
    excel_dupes = 0

    for ex in excel:
        bk = normalize(ex["rs_benef"])
        ref_k = normalize(ex["ref_interne"])
        found = None

        # 1. Exact name match
        if bk and bk in mon_by_name:
            for cand in mon_by_name[bk]:
                if cand["id"] not in used_ids:
                    found = cand
                    break
            if not found:
                found = mon_by_name[bk][0]
                excel_dupes += 1

        # 2. Ref interne match
        if not found and ref_k and ref_k in mon_by_ref:
            for cand in mon_by_ref[ref_k]:
                if cand["id"] not in used_ids:
                    found = cand
                    break

        # 3. Fuzzy name
        if not found and bk and len(bk) > 4:
            for mk, mvs in mon_by_name.items():
                if (bk in mk or mk in bk) and len(bk) > 4:
                    for cand in mvs:
                        if cand["id"] not in used_ids:
                            found = cand
                            break
                    if found: break

        if found:
            used_ids.add(found["id"])
            matched.append((ex, found))
        else:
            missing.append(ex)

    orphans = [m for m in monday if m["id"] not in used_ids]

    # Field-by-field comparison with corrected mapping
    COMPARE_FIELDS = [
        # (excel_key, monday_key, excel_normalizer, monday_normalizer, label)
        ("siren_benef", "siret_benef", norm_siren, norm_siren, "SIRET"),
        ("email_benef", "email_benef", normalize, normalize, "Email"),
        ("telephone", "telephone", norm_phone, norm_phone, "Telephone"),
        ("adresse", "adresse", normalize, normalize, "Adresse"),
        ("code_postal", "code_postal", norm_cp, norm_cp, "Code postal"),
        ("ville", "ville", normalize, normalize, "Ville"),
        ("ref_interne", "ref_interne", normalize, normalize, "Ref interne"),
        ("ref_preuve", "ref_devis", normalize, normalize, "Ref devis"),
    ]

    mismatches = []
    perfect_matches = 0
    for ex, mon in matched:
        diffs = {}
        for ex_key, mon_key, ex_norm, mon_norm, label in COMPARE_FIELDS:
            ev = ex_norm(ex.get(ex_key, ""))
            mv = mon_norm(mon.get(mon_key, ""))
            # Only flag if BOTH have values and they differ
            if ev and mv and ev != mv:
                # SIRET special: Excel has 9-digit SIREN, Monday may have 14-digit SIRET
                if label == "SIRET" and len(ev) == 9 and len(mv) == 14 and mv.startswith(ev):
                    continue  # SIREN is prefix of SIRET = OK
                diffs[label] = {"excel": ex.get(ex_key, ""), "monday": mon.get(mon_key, "")}
            elif ev and not mv:
                diffs[label] = {"excel": ex.get(ex_key, ""), "monday": "(vide)"}
            elif not ev and mv:
                pass  # Monday has data Excel doesn't = OK (enriched)
        if diffs:
            mismatches.append({
                "name": ex["rs_benef"],
                "monday_name": mon["name"],
                "monday_id": mon["id"],
                "diffs": diffs
            })
        else:
            perfect_matches += 1

    # OUTPUT
    print(f"\n{'='*60}")
    print("ECOVOLT COMPARISON — CORRECTED REPORT")
    print(f"{'='*60}")
    print(f"Excel total:         {len(excel)} (ZNI:{sc.get('ZNI',0)}, 19-09:{sc.get('19-09-2025',0)}, 01-10:{sc.get('01-10-2025',0)})")
    print(f"Monday total:        {len(monday)}")
    print(f"Matched:             {len(matched)} (dupes in Excel: {excel_dupes})")
    print(f"  Perfect matches:   {perfect_matches}")
    print(f"  With differences:  {len(mismatches)}")
    print(f"Missing from Monday: {len(missing)}")
    print(f"Orphans on Monday:   {len(orphans)}")

    print(f"\n--- ALL ORPHANS ({len(orphans)}) ---")
    for o in orphans:
        print(f"  [{o['id']}] {o['name']}")

    # Mismatch field stats
    field_counts = {}
    for m in mismatches:
        for fld in m["diffs"]:
            field_counts[fld] = field_counts.get(fld, 0) + 1
    print(f"\n--- MISMATCH BREAKDOWN BY FIELD ---")
    for fld, cnt in sorted(field_counts.items(), key=lambda x: -x[1]):
        print(f"  {fld}: {cnt} items differ")

    print(f"\n--- MISMATCHES: first 10 ---")
    for m in mismatches[:10]:
        print(f"  {m['name']} [Monday id:{m['monday_id']}]")
        for fld, vals in m["diffs"].items():
            print(f"    {fld}: Excel=\"{vals['excel'][:50]}\" vs Monday=\"{vals['monday'][:50]}\"")

    print(f"\n--- MISSING FROM MONDAY: first 15 ---")
    for m in missing[:15]:
        print(f"  [{m['sheet']}] op:{m['op_num']} {m['rs_benef']} ({m['email_benef']})")

    # Save full
    full = {
        "counts": {"excel": len(excel), "monday": len(monday), "matched": len(matched),
                    "perfect": perfect_matches, "mismatched": len(mismatches),
                    "missing": len(missing), "orphans": len(orphans), "excel_dupes": excel_dupes},
        "sheet_counts": sc,
        "orphans": [{"id": o["id"], "name": o["name"]} for o in orphans],
        "field_mismatch_counts": field_counts,
        "mismatches_all": mismatches,
        "missing_all": [{"sheet": m["sheet"], "op": m["op_num"], "benef": m["rs_benef"],
                         "email": m["email_benef"], "tel": m["telephone"]} for m in missing],
    }
    out = "/Users/john/JARVIS/projets/velo/scripts/ecovolt-comparison-v3.json"
    with open(out, "w") as f:
        json.dump(full, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nFull JSON: {out}", file=sys.stderr)

if __name__ == "__main__":
    main()

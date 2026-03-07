import openpyxl, re, json
from datetime import datetime, date

wb = openpyxl.load_workbook("/Users/john/Library/CloudStorage/Dropbox/VELO CARGO/ENEMAT/PPE - TRA-EQ-131 cloture.xlsx", data_only=True)

cutoff = datetime(2025, 9, 30)
ppe_sirets = {}  # siret -> {raison_sociale, ville, cp}
skipped_no_ppe = 0
skipped_date = 0
skipped_no_siret = 0
total_rows_checked = 0

for sn in ['01-10-2025', '19-09-2025']:
    sheet = wb[sn]
    print(f"\nProcessing sheet '{sn}': {sheet.max_row} rows")

    for row in range(2, sheet.max_row + 1):
        total_rows_checked += 1

        # Col T (20) = RAISON SOCIALE du professionnel
        col_t = sheet.cell(row=row, column=20).value
        if col_t is None:
            skipped_no_ppe += 1
            continue
        col_t_str = str(col_t).upper().strip()

        # Filter: must contain PRESERVATION or PPE
        if 'PRESERVATION' not in col_t_str and 'PPE' not in col_t_str:
            skipped_no_ppe += 1
            continue

        # Col G (7) = DATE D'ENGAGEMENT, fallback Col E (5)
        date_val = sheet.cell(row=row, column=7).value
        if date_val is None:
            date_val = sheet.cell(row=row, column=5).value

        if date_val is not None:
            if isinstance(date_val, datetime):
                d = date_val
            elif isinstance(date_val, date):
                d = datetime.combine(date_val, datetime.min.time())
            elif isinstance(date_val, str):
                try:
                    d = datetime.strptime(date_val[:10], "%Y-%m-%d")
                except:
                    d = None
            else:
                d = None

            if d is not None and d > cutoff:
                skipped_date += 1
                continue

        # Col Z (26) = SIRET
        siret_raw = sheet.cell(row=row, column=26).value
        if siret_raw is None:
            skipped_no_siret += 1
            continue

        siret_clean = re.sub(r'\D', '', str(siret_raw).split('.')[0])

        if len(siret_clean) < 9:
            skipped_no_siret += 1
            continue

        # Pad SIREN (9 digits) to SIRET (14 digits) if needed
        if len(siret_clean) == 9:
            siret_clean = siret_clean  # Keep as SIREN, we'll handle in comparison

        raison_sociale = str(sheet.cell(row=row, column=12).value or '').strip()
        ville = str(sheet.cell(row=row, column=16).value or '').strip()
        cp_raw = sheet.cell(row=row, column=15).value
        cp = str(cp_raw).split('.')[0] if cp_raw else ''

        if siret_clean not in ppe_sirets:
            ppe_sirets[siret_clean] = {
                'raison_sociale': raison_sociale,
                'ville': ville,
                'cp': cp,
                'sheet': sn
            }

print(f"\n=== EXTRACTION SUMMARY ===")
print(f"Total rows checked: {total_rows_checked}")
print(f"Skipped (no PPE/PRESERVATION): {skipped_no_ppe}")
print(f"Skipped (date > 2025-09-30): {skipped_date}")
print(f"Skipped (no valid SIRET): {skipped_no_siret}")
print(f"Unique PPE SIRETs extracted: {len(ppe_sirets)}")

# Show SIRET length distribution
lengths = {}
for s in ppe_sirets:
    l = len(s)
    lengths[l] = lengths.get(l, 0) + 1
print(f"SIRET length distribution: {lengths}")

# Save to JSON for later use
with open('/Users/john/JARVIS/enemat_ppe_sirets.json', 'w') as f:
    json.dump(ppe_sirets, f, ensure_ascii=False, indent=2)
print(f"\nSaved to /Users/john/JARVIS/enemat_ppe_sirets.json")

# Print first 10 as sample
print(f"\nSample (first 10):")
for i, (siret, info) in enumerate(list(ppe_sirets.items())[:10]):
    print(f"  {siret} | {info['raison_sociale'][:40]} | {info['ville']} | {info['cp']}")

#!/usr/bin/env python3
import json, time, sys, os, urllib.request, urllib.error
from collections import defaultdict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('ERROR: openpyxl not installed'); sys.exit(1)

API_URL = "https://api.monday.com/v2"
# Monday API token (account PPE — crm-oreka, 7 boards)
API_TOKEN = os.environ.get('MONDAY_API_KEY')
assert API_TOKEN, 'MONDAY_API_KEY env var required (PPE account)'
API_VERSION = "2024-10"
OUTPUT = "/Users/john/JARVIS/CLAUDE/Stats_PPE_Energie.xlsx"
NAF_FILE = "/Users/john/JARVIS/projets/velo/docs/naf-enemat-reference.json"

BOARDS = [
    {"name": "ATHOME",   "id": 2144986053, "naf_col": "text_mkvkhf8p", "val_naf_col": "color_mm0vrhdk"},
    {"name": "JM",       "id": 2137662048, "naf_col": "text_mkvkj1mb", "val_naf_col": "color_mm0vgkmf"},
    {"name": "SALIH",    "id": 5013455904, "naf_col": "text_mkvk64jp", "val_naf_col": "color_mm0vf9d2"},
    {"name": "STELLARS", "id": 5001072451, "naf_col": "text_mkvk64jp", "val_naf_col": "color_mm0vq45a"},
    {"name": "ALEX",     "id": 5002798369, "naf_col": "text_mkvk64jp", "val_naf_col": "color_mm0vz342"},
    {"name": "EKL",      "id": 2140187165, "naf_col": "text_mkvks2a4", "val_naf_col": "color_mm0v8049"},
    {"name": "DIZIEN",   "id": 2146667697, "naf_col": "text_mkvk64jp", "val_naf_col": "color_mm0v75kg"},
]

DEPOT_MAP = {
    "Ile-de-France": ["75","77","78","91","92","93","94","95"],
    "Lyon": ["01","07","26","38","42","43","63","69","73","74"],
    "Nantes": ["22","29","35","44","49","53","56","72","85"],
    "Bordeaux": ["16","17","19","23","24","33","40","47","64","79","86","87"],
    "Marseille": ["04","05","06","13","20","30","34","48","66","83","84"],
}
P2D = {}
for d, ps in DEPOT_MAP.items():
    for p in ps:
        P2D[p] = d

FD = Font(name="Calibri", size=11)
FT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
FS = Font(name="Calibri", size=11, italic=True, color="555555")
FH = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FO = Font(name="Calibri", size=11, bold=True)
PH = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
PA = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
PT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PW = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
TB = Border(left=Side(style="thin",color="B0B0B0"),right=Side(style="thin",color="B0B0B0"),
            top=Side(style="thin",color="B0B0B0"),bottom=Side(style="thin",color="B0B0B0"))


def mquery(query):
    payload = json.dumps({"query": query}).encode("utf-8")
    req = urllib.request.Request(API_URL, data=payload,
        headers={"Authorization": API_TOKEN, "Content-Type": "application/json", "API-Version": API_VERSION}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            if "errors" in data:
                print(f"  API ERRORS: {json.dumps(data['errors'])[:200]}")
            return data.get("data", {})
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8") if e.fp else ""
        print(f"  HTTP Error {e.code}: {body[:300]}"); return {}
    except Exception as e:
        print(f"  Request error: {e}"); return {}


def fetch_items(board):
    bid = board["id"]; nc = board["naf_col"]; vc = board["val_naf_col"]
    cols = ["status","numeric_mkvj879j","numeric_mkvj6e60","numeric_mkvjbazm","text_mkvjgcp9",nc,vc]
    cs = ", ".join(f'"{c}"' for c in cols)
    all_items = []
    q = '{boards(ids: [%d]) {items_page(limit: 500) {cursor items {id name column_values(ids: [%s]) {id text value}}}}}' % (bid, cs)
    data = mquery(q)
    if not data or "boards" not in data or not data["boards"]:
        print(f"  ERROR: No data for {board['name']}"); return []
    page = data["boards"][0].get("items_page", {})
    items = page.get("items", [])
    cursor = page.get("cursor")
    all_items.extend(items)
    print(f"  Page 1: {len(items)} items (cursor: {'yes' if cursor else 'no'})")
    pn = 1
    while cursor:
        time.sleep(0.5); pn += 1
        ce = cursor.replace('"', '\\"')
        q = '{next_items_page(limit: 500, cursor: "%s") {cursor items {id name column_values(ids: [%s]) {id text value}}}}' % (ce, cs)
        data = mquery(q)
        if not data or "next_items_page" not in data:
            print(f"  ERROR: Pagination failed page {pn}"); break
        page = data["next_items_page"]
        items = page.get("items", [])
        cursor = page.get("cursor")
        all_items.extend(items)
        print(f"  Page {pn}: {len(items)} items (cursor: {'yes' if cursor else 'no'})")
    return all_items


def parse_items(items, board):
    records = []; nc = board["naf_col"]; vc = board["val_naf_col"]
    for item in items:
        cols = {cv["id"]: cv for cv in item.get("column_values", [])}
        statut = (cols.get("status", {}).get("text") or "").strip()
        vvs = (cols.get("numeric_mkvj6e60", {}).get("text") or "").strip()
        vws = (cols.get("numeric_mkvj879j", {}).get("text") or "").strip()
        try: vv = int(float(vvs)) if vvs else 0
        except: vv = 0
        try: vw = int(float(vws)) if vws else 0
        except: vw = 0
        velos = vv if vv > 0 else vw
        naf = (cols.get(nc, {}).get("text") or "").strip()
        vnt = (cols.get(vc, {}).get("text") or "").strip().upper()
        if vnt in ("OUI", "OK"): vnaf = "OUI"
        elif vnt in ("NON", "KO"): vnaf = "NON"
        else: vnaf = "A VERIFIER"
        cpr = (cols.get("numeric_mkvjbazm", {}).get("text") or "").strip()
        try: cp = str(int(float(cpr))).zfill(5) if cpr else ""
        except: cp = cpr
        ville = (cols.get("text_mkvjgcp9", {}).get("text") or "").strip()
        records.append({"id": item["id"], "name": item.get("name",""), "board": board["name"],
            "statut": statut, "velos": velos, "naf_code": naf, "validation_naf": vnaf,
            "cp": cp, "ville": ville, "is_cv": statut.upper().startswith("CONTROL VALID"),
            "is_ok": vnaf == "OUI", "is_ko": vnaf == "NON",
            "has_1v": velos == 1, "has_mv": velos > 1})
    return records


def cstats(recs):
    tc = len(recs); tv = sum(r["velos"] for r in recs)
    cok = sum(1 for r in recs if r["is_ok"]); cko = sum(1 for r in recs if r["is_ko"])
    vok = sum(r["velos"] for r in recs if r["is_ok"]); vko = sum(r["velos"] for r in recs if r["is_ko"])
    c1ok = sum(1 for r in recs if r["has_1v"] and r["is_ok"])
    c1ko = sum(1 for r in recs if r["has_1v"] and r["is_ko"])
    cmok = sum(1 for r in recs if r["has_mv"] and r["is_ok"])
    cmko = sum(1 for r in recs if r["has_mv"] and r["is_ko"])
    v1ok = sum(r["velos"] for r in recs if r["has_1v"] and r["is_ok"])
    v1ko = sum(r["velos"] for r in recs if r["has_1v"] and r["is_ko"])
    vmok = sum(r["velos"] for r in recs if r["has_mv"] and r["is_ok"])
    vmko = sum(r["velos"] for r in recs if r["has_mv"] and r["is_ko"])
    return {"tc":tc,"tv":tv,"cok":cok,"cko":cko,"vok":vok,"vko":vko,
            "c1ok":c1ok,"c1ko":c1ko,"cmok":cmok,"cmko":cmko,
            "v1ok":v1ok,"v1ko":v1ko,"vmok":vmok,"vmko":vmko}


def naf_ko_detail(recs, labels):
    nd = defaultdict(lambda: {"c":0,"v":0,"c1":0,"cm":0,"v1":0,"vm":0})
    for r in recs:
        if not r["is_ko"]: continue
        code = r["naf_code"] or "(vide)"
        nd[code]["c"] += 1; nd[code]["v"] += r["velos"]
        if r["has_1v"]: nd[code]["c1"] += 1; nd[code]["v1"] += r["velos"]
        if r["has_mv"]: nd[code]["cm"] += 1; nd[code]["vm"] += r["velos"]
    res = []
    for code, d in nd.items():
        res.append({"code":code,"label":labels.get(code,""),"c":d["c"],"v":d["v"],
                     "c1":d["c1"],"cm":d["cm"],"v1":d["v1"],"vm":d["vm"]})
    res.sort(key=lambda x: x["v"], reverse=True)
    return res


def depot_stats(recs):
    ds = defaultdict(lambda: {"ct":0,"co":0,"ck":0,"vt":0,"vo":0,"vk":0})
    for r in recs:
        pfx = r["cp"][:2] if len(r["cp"]) >= 2 else ""
        dn = P2D.get(pfx, "Autre")
        ds[dn]["ct"] += 1; ds[dn]["vt"] += r["velos"]
        if r["is_ok"]: ds[dn]["co"] += 1; ds[dn]["vo"] += r["velos"]
        if r["is_ko"]: ds[dn]["ck"] += 1; ds[dn]["vk"] += r["velos"]
    order = ["Ile-de-France","Lyon","Nantes","Bordeaux","Marseille","Autre"]
    res = []
    for n in order:
        if n in ds: ds[n]["depot"] = n; res.append(ds[n])
    for n in ds:
        if n not in order: ds[n]["depot"] = n; res.append(ds[n])
    return res


def pct(a, b):
    return "0.0%" if b == 0 else f"{a/b*100:.1f}%"


def aw(ws, mn=10, mx=50):
    for cc in ws.columns:
        ml = mn; cl = get_column_letter(cc[0].column)
        for c in cc:
            if c.value:
                l = len(str(c.value))
                if l > ml: ml = l
        ws.column_dimensions[cl].width = min(ml + 3, mx)


def shr(ws, row, c1, c2):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = FH; cell.fill = PH; cell.alignment = AC; cell.border = TB


def sdr(ws, row, c1, c2, alt=False, tot=False):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.font = FO if tot else FD
        if tot: cell.fill = PT
        elif alt: cell.fill = PA
        else: cell.fill = PW
        cell.border = TB
        if isinstance(cell.value, (int, float)): cell.alignment = AR
        elif isinstance(cell.value, str) and cell.value.endswith("%"): cell.alignment = AC
        else: cell.alignment = AL


def wst(ws, sr, title, s):
    row = sr
    ws.cell(row=row, column=1, value=title).font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    row += 1
    hdrs = ["Indicateur","Total","OK (NAF eligible)","KO (NAF non eligible)","% OK","% KO"]
    for i, h in enumerate(hdrs, 1): ws.cell(row=row, column=i, value=h)
    shr(ws, row, 1, len(hdrs)); row += 1
    rows = [
        ("Nombre total de clients", s["tc"], s["cok"], s["cko"], pct(s["cok"],s["tc"]), pct(s["cko"],s["tc"])),
        ("Nombre total de velos", s["tv"], s["vok"], s["vko"], pct(s["vok"],s["tv"]), pct(s["vko"],s["tv"])),
        ("Clients avec 1 velo", s["c1ok"]+s["c1ko"], s["c1ok"], s["c1ko"], pct(s["c1ok"],s["c1ok"]+s["c1ko"]), pct(s["c1ko"],s["c1ok"]+s["c1ko"])),
        ("Clients avec +1 velo", s["cmok"]+s["cmko"], s["cmok"], s["cmko"], pct(s["cmok"],s["cmok"]+s["cmko"]), pct(s["cmko"],s["cmok"]+s["cmko"])),
        ("Velos (clients 1 velo)", s["v1ok"]+s["v1ko"], s["v1ok"], s["v1ko"], pct(s["v1ok"],s["v1ok"]+s["v1ko"]), pct(s["v1ko"],s["v1ok"]+s["v1ko"])),
        ("Velos (clients +1 velo)", s["vmok"]+s["vmko"], s["vmok"], s["vmko"], pct(s["vmok"],s["vmok"]+s["vmko"]), pct(s["vmko"],s["vmok"]+s["vmko"])),
    ]
    for idx, dr in enumerate(rows):
        for i, v in enumerate(dr, 1): ws.cell(row=row, column=i, value=v)
        sdr(ws, row, 1, len(hdrs), alt=(idx%2==1)); row += 1
    return row + 1


def wnaf(ws, title, subtitle, nklist):
    row = 1
    ws.cell(row=row, column=1, value=title).font = FT; row += 1
    ws.cell(row=row, column=1, value=subtitle).font = FS; row += 2
    hdrs = ["Code NAF","Libelle activite","Clients KO","Velos KO","Clients KO (1 velo)","Clients KO (+1 velo)","Velos KO (1 velo)","Velos KO (+1 velo)"]
    for i, h in enumerate(hdrs, 1): ws.cell(row=row, column=i, value=h)
    shr(ws, row, 1, len(hdrs)); row += 1
    tc=tv=t1=tm=tv1=tvm=0
    for idx, n in enumerate(nklist):
        vals = [n["code"],n["label"],n["c"],n["v"],n["c1"],n["cm"],n["v1"],n["vm"]]
        for i, v in enumerate(vals, 1): ws.cell(row=row, column=i, value=v)
        sdr(ws, row, 1, len(hdrs), alt=(idx%2==1))
        tc+=n["c"];tv+=n["v"];t1+=n["c1"];tm+=n["cm"];tv1+=n["v1"];tvm+=n["vm"]; row+=1
    tots = ["TOTAL","",tc,tv,t1,tm,tv1,tvm]
    for i, v in enumerate(tots, 1): ws.cell(row=row, column=i, value=v)
    sdr(ws, row, 1, len(hdrs), tot=True)
    ws.freeze_panes = "A5"; aw(ws)


def wdep(ws, title, subtitle, dlist):
    row = 1
    ws.cell(row=row, column=1, value=title).font = FT; row += 1
    ws.cell(row=row, column=1, value=subtitle).font = FS; row += 2
    hdrs = ["Depot","Clients Total","Clients OK","Clients KO","% Clients OK","Velos Total","Velos OK","Velos KO","% Velos OK"]
    for i, h in enumerate(hdrs, 1): ws.cell(row=row, column=i, value=h)
    shr(ws, row, 1, len(hdrs)); row += 1
    tc=tv=tco=tck=tvo=tvk=0
    for idx, d in enumerate(dlist):
        vals = [d["depot"],d["ct"],d["co"],d["ck"],pct(d["co"],d["ct"]),d["vt"],d["vo"],d["vk"],pct(d["vo"],d["vt"])]
        for i, v in enumerate(vals, 1): ws.cell(row=row, column=i, value=v)
        sdr(ws, row, 1, len(hdrs), alt=(idx%2==1))
        tc+=d["ct"];tco+=d["co"];tck+=d["ck"];tv+=d["vt"];tvo+=d["vo"];tvk+=d["vk"]; row+=1
    tots = ["TOTAL",tc,tco,tck,pct(tco,tc),tv,tvo,tvk,pct(tvo,tv)]
    for i, v in enumerate(tots, 1): ws.cell(row=row, column=i, value=v)
    sdr(ws, row, 1, len(hdrs), tot=True)
    ws.freeze_panes = "A5"; aw(ws)


def main():
    print("="*60)
    print("PPE ENERGIE - Extraction Monday.com + Stats Excel")
    print("="*60)
    print("\n[1/3] Chargement reference NAF...")
    try:
        with open(NAF_FILE, "r", encoding="utf-8") as f:
            naf_labels = json.load(f).get("labels", {})
        print(f"  {len(naf_labels)} codes NAF charges")
    except Exception as e:
        print(f"  ERREUR: {e}"); naf_labels = {}

    print("\n[2/3] Extraction des 7 boards PPE...")
    all_recs = []; brecs = {}
    for b in BOARDS:
        print(f"\n  --- Board: {b['name']} (ID: {b['id']}) ---")
        time.sleep(0.5)
        items = fetch_items(b)
        recs = parse_items(items, b)
        brecs[b["name"]] = recs; all_recs.extend(recs)
        print(f"  => {len(recs)} items extraits")
    print(f"\n  TOTAL: {len(all_recs)} items sur 7 boards")

    cv = [r for r in all_recs if r["is_cv"]]
    ot = [r for r in all_recs if not r["is_cv"]]
    print(f"  CONTROLE VALIDE: {len(cv)} | EN ATTENTE: {len(ot)}")

    print("\n[3/3] Calcul stats + generation Excel...")
    scv = cstats(cv); sot = cstats(ot)
    bscv = {}
    for b in BOARDS:
        br = [r for r in brecs[b["name"]] if r["is_cv"]]
        bscv[b["name"]] = cstats(br)
    nkcv = naf_ko_detail(cv, naf_labels); nkot = naf_ko_detail(ot, naf_labels)
    dcv = depot_stats(cv); dot = depot_stats(ot)

    wb = openpyxl.Workbook()
    # Onglet 1: SYNTHESE
    ws1 = wb.active; ws1.title = "SYNTHESE"; ws1.sheet_properties.tabColor = "1F4E79"
    row = 1
    ws1.cell(row=row, column=1, value="PPE ENERGIE - Statistiques Validation NAF ENEMAT").font = FT; row += 1
    ws1.cell(row=row, column=1, value="Date : 23/02/2026 | Perimetre : 7 boards PPE Energie").font = FS; row += 2
    row = wst(ws1, row, "Clients CONTROLE VALIDE", scv); row += 1
    row = wst(ws1, row, "Clients hors CONTROLE VALIDE (en attente)", sot)
    ws1.freeze_panes = "A4"; aw(ws1)

    # Onglet 2: DETAIL PAR BOARD
    ws2 = wb.create_sheet("DETAIL PAR BOARD"); ws2.sheet_properties.tabColor = "00B050"
    row = 1
    ws2.cell(row=row, column=1, value="PPE ENERGIE - Detail par Board (CONTROLE VALIDE)").font = FT; row += 1
    ws2.cell(row=row, column=1, value="Date : 23/02/2026 | Perimetre : Clients CONTROLE VALIDE uniquement").font = FS; row += 2
    for b in BOARDS:
        bn = b["name"]; bs = bscv[bn]
        tcv = len([r for r in brecs[bn] if r["is_cv"]])
        row = wst(ws2, row, f"Board {bn} ({tcv} clients CV)", bs); row += 1
    ws2.freeze_panes = "A4"; aw(ws2)

    # Onglet 3: NAF KO CV
    ws3 = wb.create_sheet("NAF KO - CONTROLE VALIDE"); ws3.sheet_properties.tabColor = "FF0000"
    wnaf(ws3, "Codes NAF Non Eligibles - Clients CONTROLE VALIDE",
         f"Date : 23/02/2026 | {len(nkcv)} codes NAF KO distincts | {scv['cko']} clients | {scv['vko']} velos", nkcv)

    # Onglet 4: NAF KO EN ATTENTE
    ws4 = wb.create_sheet("NAF KO - EN ATTENTE"); ws4.sheet_properties.tabColor = "FFC000"
    wnaf(ws4, "Codes NAF Non Eligibles - Clients EN ATTENTE",
         f"Date : 23/02/2026 | {len(nkot)} codes NAF KO distincts | {sot['cko']} clients | {sot['vko']} velos", nkot)

    # Onglet 5: DEPOTS CV
    ws5 = wb.create_sheet("DEPOTS - CONTROLE VALIDE"); ws5.sheet_properties.tabColor = "7030A0"
    wdep(ws5, "Ventilation Geographique - Clients CONTROLE VALIDE", "Date : 23/02/2026 | Repartition par depot logistique", dcv)

    # Onglet 6: DEPOTS EN ATTENTE
    ws6 = wb.create_sheet("DEPOTS - EN ATTENTE"); ws6.sheet_properties.tabColor = "808080"
    wdep(ws6, "Ventilation Geographique - Clients EN ATTENTE", "Date : 23/02/2026 | Repartition par depot logistique", dot)

    wb.save(OUTPUT)
    print(f"\n  Fichier Excel sauvegarde : {OUTPUT}")

    print("\n" + "="*60)
    print("RESUME FINAL")
    print("="*60)
    print(f"\n{'Board':<12} {'Total':>7} {'CV':>7} {'Autre':>7}")
    print("-"*40)
    for b in BOARDS:
        bn = b["name"]; t = len(brecs[bn]); c = len([r for r in brecs[bn] if r["is_cv"]]); o = t - c
        print(f"{bn:<12} {t:>7} {c:>7} {o:>7}")
    print("-"*40)
    print(f"{'TOTAL':<12} {len(all_recs):>7} {len(cv):>7} {len(ot):>7}")
    print(f"\n--- CONTROLE VALIDE ---")
    print(f"  Clients: {scv['tc']} | Velos: {scv['tv']}")
    print(f"  OK: {scv['cok']} clients / {scv['vok']} velos")
    print(f"  KO: {scv['cko']} clients / {scv['vko']} velos")
    print(f"  NAF KO distincts: {len(nkcv)}")
    print(f"\n--- EN ATTENTE ---")
    print(f"  Clients: {sot['tc']} | Velos: {sot['tv']}")
    print(f"  OK: {sot['cok']} clients / {sot['vok']} velos")
    print(f"  KO: {sot['cko']} clients / {sot['vko']} velos")
    print(f"  NAF KO distincts: {len(nkot)}")
    print("\nDone.")

if __name__ == "__main__":
    main()

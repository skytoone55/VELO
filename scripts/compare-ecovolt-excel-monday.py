#!/usr/bin/env python3
"""
Compare Ecovolt entries from Excel reference vs Monday.com board.
Output: compact JSON report to stdout.
"""
import json, sys, os, re, time
import openpyxl, requests

EXCEL_PATH = "/Users/john/JARVIS/projets/velo/docs/PPE - TRA-EQ-131 cloture-complet.xlsx"
BOARD_ID = 9990833105

# Read API key
with open("/Users/john/JARVIS/projets/velo/.env.local") as f:
    for line in f:
        if line.startswith("MONDAY_API_KEY"):
            API_KEY = line.split("=", 1)[1].strip().strip('"')
            break

MONDAY_URL = "https://api.monday.com/v2"
HEADERS = {"Authorization": API_KEY, "Content-Type": "application/json"}

###############################################################################
# 1. EXTRACT ECOVOLT FROM EXCEL
###############################################################################
def extract_ecovolt_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    all_entries = []
    sheet_counts = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheet_counts[sheet_name] = 0
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Find key column indices
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower().replace("\n", " ").replace("\r", " ")
            if "raison sociale" in hl and "professionnel" in hl:
                col_map["raison_sociale_pro"] = i
            elif "raison sociale" in hl and "bénéficiaire" in hl:
                col_map["raison_sociale_benef"] = i
            elif "raison sociale" in hl and "demandeur" in hl:
                col_map["raison_sociale_demandeur"] = i
            elif "opération n" in hl or hl.startswith("opération n"):
                col_map["operation_num"] = i
            elif "siren" in hl and "professionnel" in hl:
                col_map["siren_pro"] = i
            elif "siren" in hl and "bénéficiaire" in hl:
                col_map["siren_benef"] = i
            elif "reference interne" in hl:
                col_map["ref_interne"] = i
            elif "adresse" in hl and "siège" in hl:
                col_map["adresse"] = i
            elif "code postal" in hl:
                col_map["code_postal"] = i
            elif "ville" in hl:
                col_map["ville"] = i
            elif "téléphone" in hl or "telephone" in hl:
                col_map["telephone"] = i
            elif "courriel" in hl or "email" in hl:
                col_map["email"] = i
            elif "reference" in hl and "preuve" in hl:
                col_map["ref_preuve"] = i

        count = 0
        for row in rows[1:]:
            # Check if this row is Ecovolt
            rs_pro_idx = col_map.get("raison_sociale_pro")
            if rs_pro_idx is None or rs_pro_idx >= len(row):
                continue
            rs_pro = str(row[rs_pro_idx] or "").strip()
            if "eco-volt" not in rs_pro.lower() and "ecovolt" not in rs_pro.lower():
                continue

            def get_val(key):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    if v is None:
                        return ""
                    return str(v).strip()
                return ""

            entry = {
                "sheet": sheet_name,
                "operation_num": get_val("operation_num"),
                "raison_sociale_pro": rs_pro,
                "siren_pro": get_val("siren_pro"),
                "raison_sociale_benef": get_val("raison_sociale_benef"),
                "siren_benef": get_val("siren_benef"),
                "ref_interne": get_val("ref_interne"),
                "adresse": get_val("adresse"),
                "code_postal": get_val("code_postal"),
                "ville": get_val("ville"),
                "telephone": get_val("telephone"),
                "email": get_val("email"),
                "ref_preuve": get_val("ref_preuve"),
            }
            all_entries.append(entry)
            count += 1

        sheet_counts[sheet_name] = count

    wb.close()
    return all_entries, sheet_counts

###############################################################################
# 2. FETCH ALL MONDAY ITEMS
###############################################################################
def fetch_monday_items():
    all_items = []

    # First page
    query = '''{ boards(ids: %d) { items_page(limit: 500) { cursor items { id name column_values { id text value } } } } }''' % BOARD_ID
    resp = requests.post(MONDAY_URL, headers=HEADERS, json={"query": query}, timeout=30)
    data = resp.json()

    page = data["data"]["boards"][0]["items_page"]
    all_items.extend(page["items"])
    cursor = page["cursor"]

    while cursor:
        query2 = '''{ next_items_page(limit: 500, cursor: "%s") { cursor items { id name column_values { id text value } } } }''' % cursor
        resp = requests.post(MONDAY_URL, headers=HEADERS, json={"query": query2}, timeout=30)
        data = resp.json()
        page2 = data["data"]["next_items_page"]
        all_items.extend(page2["items"])
        cursor = page2["cursor"]
        time.sleep(0.3)  # rate limit

    return all_items

###############################################################################
# 3. COMPARE
###############################################################################
def normalize(s):
    """Normalize string for comparison: lowercase, strip, collapse whitespace."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    # Remove trailing .0 for numbers
    s = re.sub(r'\.0$', '', s)
    return s

def monday_item_to_dict(item):
    """Extract relevant fields from Monday item column_values."""
    d = {"id": item["id"], "name": item["name"]}
    for cv in item["column_values"]:
        cid = cv["id"]
        txt = cv.get("text") or ""
        d[cid] = txt.strip()
    return d

def compare_entries():
    print("Extracting Ecovolt from Excel...", file=sys.stderr)
    excel_entries, sheet_counts = extract_ecovolt_from_excel()
    print(f"  Found {len(excel_entries)} entries: {sheet_counts}", file=sys.stderr)

    print("Fetching Monday items...", file=sys.stderr)
    monday_raw = fetch_monday_items()
    print(f"  Found {len(monday_raw)} items", file=sys.stderr)

    monday_items = [monday_item_to_dict(item) for item in monday_raw]

    # Build Monday lookup by normalized name
    monday_by_name = {}
    for mi in monday_items:
        key = normalize(mi["name"])
        if key in monday_by_name:
            monday_by_name[key].append(mi)
        else:
            monday_by_name[key] = [mi]

    # Build Excel lookup by raison_sociale_benef (this is what Monday item name should be)
    # Actually, Monday item name = "Nom de l'opération" which in Excel is raison_sociale_benef
    # Let's also try operation_num as fallback

    matched = []
    mismatched = []
    missing_from_monday = []
    matched_monday_ids = set()

    for ex in excel_entries:
        # Try matching by beneficiary name (most common)
        benef_key = normalize(ex["raison_sociale_benef"])
        op_num_key = normalize(ex["operation_num"])
        ref_key = normalize(ex["ref_interne"])

        found = None
        match_method = ""

        # Try benef name
        if benef_key and benef_key in monday_by_name:
            candidates = monday_by_name[benef_key]
            found = candidates[0]
            match_method = "benef_name"

        if not found:
            # Try op number as name
            if op_num_key and op_num_key in monday_by_name:
                found = monday_by_name[op_num_key][0]
                match_method = "op_num"

        if not found:
            # Fuzzy: check if benef name is contained in any Monday name or vice versa
            if benef_key and len(benef_key) > 3:
                for mk, mvs in monday_by_name.items():
                    if benef_key in mk or mk in benef_key:
                        found = mvs[0]
                        match_method = "fuzzy_benef"
                        break

        if found:
            matched_monday_ids.add(found["id"])
            matched.append({"excel": ex, "monday": found, "method": match_method})
        else:
            missing_from_monday.append(ex)

    # Find orphans on Monday (not matched to any Excel entry)
    orphans = [mi for mi in monday_items if mi["id"] not in matched_monday_ids]

    # Now compare matched entries field by field
    # We need to discover which Monday column IDs map to which fields
    # Let's dump a sample Monday item to see column IDs
    sample_monday = monday_items[0] if monday_items else {}
    monday_col_ids = {k: v for k, v in sample_monday.items() if k not in ("id", "name")}

    # Build report
    report = {
        "counts": {
            "excel_total": len(excel_entries),
            "excel_by_sheet": sheet_counts,
            "monday_total": len(monday_items),
            "matched": len(matched),
            "missing_from_monday": len(missing_from_monday),
            "orphans_on_monday": len(orphans),
        },
        "monday_column_ids_sample": monday_col_ids,
        "orphans": [{"name": o["name"], "id": o["id"]} for o in orphans[:50]],
        "missing_first_20": [
            {"benef": e["raison_sociale_benef"], "op": e["operation_num"],
             "sheet": e["sheet"], "email": e["email"]}
            for e in missing_from_monday[:20]
        ],
        "matched_sample_5": [
            {"excel_benef": m["excel"]["raison_sociale_benef"],
             "monday_name": m["monday"]["name"],
             "method": m["method"],
             "monday_id": m["monday"]["id"]}
            for m in matched[:5]
        ],
    }

    return report

if __name__ == "__main__":
    report = compare_entries()
    # Write report to file
    out_path = "/Users/john/JARVIS/projets/velo/scripts/ecovolt-comparison-report.json"
    with open(out_path, "w") as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)
    print(f"Report written to {out_path}", file=sys.stderr)

    # Print compact summary to stdout
    c = report["counts"]
    print(f"\n{'='*60}")
    print(f"ECOVOLT COMPARISON REPORT")
    print(f"{'='*60}")
    print(f"Excel Ecovolt entries: {c['excel_total']} ({c['excel_by_sheet']})")
    print(f"Monday items:          {c['monday_total']}")
    print(f"Matched:               {c['matched']}")
    print(f"Missing from Monday:   {c['missing_from_monday']}")
    print(f"Orphans on Monday:     {c['orphans_on_monday']}")
    print(f"\n--- ORPHANS (Monday items not in Excel) ---")
    for o in report["orphans"]:
        print(f"  [{o['id']}] {o['name']}")
    print(f"\n--- MISSING FROM MONDAY (first 20) ---")
    for m in report["missing_first_20"]:
        print(f"  [{m['sheet']}] {m['benef']} (op:{m['op']}, email:{m['email']})")
    print(f"\n--- MATCHED SAMPLE (first 5) ---")
    for m in report["matched_sample_5"]:
        print(f"  Excel: {m['excel_benef']} -> Monday: {m['monday_name']} (id:{m['monday_id']}, method:{m['method']})")

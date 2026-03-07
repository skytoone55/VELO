import openpyxl
from datetime import datetime

# --- Read ENEMAT PPE SIRETs ---
wb1 = openpyxl.load_workbook(
    "/Users/john/Library/CloudStorage/Dropbox/VELO CARGO/ENEMAT/PPE - TRA-EQ-131 cloture.xlsx",
    read_only=True, data_only=True
)
cutoff = datetime(2025, 9, 30, 23, 59, 59)
enemat_sirets = set()
enemat_details = {}

for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    rows_list = list(ws.iter_rows(min_row=2, values_only=True))
    for r in rows_list:
        if not r or len(r) <= 25:
            continue
        col_t = str(r[19] or '').upper() if r[19] else ''
        if 'PRESERVATION' not in col_t and 'PPE' not in col_t:
            continue
        date_g = r[6] if len(r) > 6 else None
        date_e = r[4] if len(r) > 4 else None
        ref_date = date_g or date_e
        if ref_date and isinstance(ref_date, datetime) and ref_date > cutoff:
            continue
        siret_raw = r[25]
        if not siret_raw:
            continue
        if isinstance(siret_raw, (int, float)):
            siret = str(int(siret_raw)).strip()
        else:
            siret = str(siret_raw).strip()
        siret = siret.zfill(14)
        if len(siret) >= 14:
            enemat_sirets.add(siret)
            if siret not in enemat_details:
                enemat_details[siret] = {
                    'raison_soc': str(r[11] or '') if len(r) > 11 and r[11] else '',
                    'ville': str(r[15] or '') if len(r) > 15 and r[15] else '',
                }

wb1.close()
print(f"ENEMAT PPE (<=30/09/2025): {len(enemat_sirets)} SIRETs uniques")

# --- Read Internal Excel SIRETs ---
wb2 = openpyxl.load_workbook(
    "/Users/john/JARVIS/velo/donnee excel.xlsx",
    read_only=True, data_only=True
)
ws2 = wb2[wb2.sheetnames[0]]
internal_sirets = set()
internal_details = {}

for r in ws2.iter_rows(min_row=2, values_only=True):
    if not r or len(r) <= 25:
        continue
    siret_raw = r[25]
    if not siret_raw:
        continue
    if isinstance(siret_raw, (int, float)):
        siret = str(int(siret_raw)).strip()
    else:
        siret = str(siret_raw).strip()
    siret = siret.zfill(14)
    if len(siret) >= 14:
        internal_sirets.add(siret)
        if siret not in internal_details:
            internal_details[siret] = {
                'raison_soc': str(r[13] or '') if len(r) > 13 and r[13] else '',
                'ville': str(r[28] or '') if len(r) > 28 and r[28] else '',
            }

wb2.close()
print(f"Excel interne: {len(internal_sirets)} SIRETs uniques")

# --- Compare ---
in_both = enemat_sirets & internal_sirets
in_enemat_only = enemat_sirets - internal_sirets
in_internal_only = internal_sirets - enemat_sirets

print(f"\n{'='*55}")
print(f"  COMPARAISON PAR SIRET (14 chiffres)")
print(f"{'='*55}\n")
print(f"  ENEMAT PPE (<=30/09/2025) : {len(enemat_sirets):>5} SIRETs uniques")
print(f"  Excel interne             : {len(internal_sirets):>5} SIRETs uniques")
print(f"")
print(f"  Presents dans les DEUX    : {len(in_both):>5}")
print(f"  Dans ENEMAT seulement     : {len(in_enemat_only):>5}")
print(f"  Dans Excel interne seul   : {len(in_internal_only):>5}")

print(f"\n{'='*55}")
print(f"  CLIENTS ENEMAT ABSENTS DE L'EXCEL INTERNE ({len(in_enemat_only)})")
print(f"{'='*55}\n")
for i, siret in enumerate(sorted(in_enemat_only), 1):
    d = enemat_details.get(siret, {})
    rs = d.get('raison_soc', '')
    vi = d.get('ville', '')
    print(f"  {i:>3}. SIRET {siret}  {rs}  [{vi}]")

print(f"\n{'='*55}")
print(f"  CLIENTS EXCEL INTERNE ABSENTS D'ENEMAT ({len(in_internal_only)})")
print(f"{'='*55}\n")
count = 0
for siret in sorted(in_internal_only):
    count += 1
    d = internal_details.get(siret, {})
    rs = d.get('raison_soc', '')
    vi = d.get('ville', '')
    if count <= 20:
        print(f"  {count:>3}. SIRET {siret}  {rs}  [{vi}]")
if count > 20:
    print(f"\n  ... et {count - 20} autres")

print(f"\n{'='*55}")
print(f"  FIN DE LA COMPARAISON")
print(f"{'='*55}")

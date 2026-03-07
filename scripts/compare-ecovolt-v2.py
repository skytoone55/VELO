#!/usr/bin/env python3
"""
V2: Field-by-field comparison of matched Ecovolt entries.
Monday column mapping:
  text_mkvfykn9  = SIRET (or SIREN) professionnel
  text_mkvf8zp6  = Reference retina / dossier
  email_mkvfk63f = Email beneficiaire
  long_text_mkvn5k9w = Telephone beneficiaire (stored as long_text!)
  text_mkvfetg2  = Adresse siege social
  text_mkvfhcn9  = Code postal
  text_mkvfgh8t  = Ville
  text_mkvft2w3  = Code NAF/APE
  email_mkvfnv4q = Email (2nd? or contact)
"""
import json, sys, re, time
import openpyxl, requests

EXCEL_PATH = "/Users/john/JARVIS/projets/velo/docs/PPE - TRA-EQ-131 cloture-complet.xlsx"
BOARD_ID = 9990833105

with open("/Users/john/JARVIS/projets/velo/.env.local") as f:
    for line in f:
        if line.startswith("MONDAY_API_KEY"):
            API_KEY = line.split("=", 1)[1].strip().strip('"')
            break

MONDAY_URL = "https://api.monday.com/v2"
HEADERS = {"Authorization": API_KEY, "Content-Type": "application/json"}

# Monday column ID -> field name mapping
MONDAY_COL_MAP = {
    "text_mkvfykn9": "siren_pro",      # SIRET/SIREN professionnel
    "text_mkvf8zp6": "ref_dossier",     # Ref retina dossier
    "email_mkvfk63f": "email",          # Email beneficiaire
    "long_text_mkvn5k9w": "telephone",  # Telephone
    "text_mkvfetg2": "adresse",         # Adresse
    "text_mkvfhcn9": "code_postal",     # Code postal
    "text_mkvfgh8t": "ville",           # Ville
    "text_mkvft2w3": "naf",             # Code NAF
}

def normalize(s):
    if not s: return ""
    s = str(s).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'\.0$', '', s)
    return s

def norm_phone(s):
    if not s: return ""
    return re.sub(r'[^0-9+]', '', str(s))

def norm_cp(s):
    if not s: return ""
    s = str(s).strip()
    s = re.sub(r'\.0$', '', s)
    return s.zfill(5) if s.isdigit() and len(s) < 5 else s

def extract_ecovolt():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    entries = []
    sheet_counts = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheet_counts[sn] = 0
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower().replace("\n"," ").replace("\r"," ")
            if "raison sociale" in hl and "professionnel" in hl: col_map["rs_pro"] = i
            elif "raison sociale" in hl and "bénéficiaire" in hl: col_map["rs_benef"] = i
            elif "opération n" in hl: col_map["op_num"] = i
            elif "siren" in hl and "professionnel" in hl: col_map["siren_pro"] = i
            elif "reference interne" in hl: col_map["ref_interne"] = i
            elif "adresse" in hl and "siège" in hl: col_map["adresse"] = i
            elif "code postal" in hl: col_map["code_postal"] = i
            elif "ville" in hl: col_map["ville"] = i
            elif "téléphone" in hl or "telephone" in hl: col_map["telephone"] = i
            elif "courriel" in hl: col_map["email"] = i
            elif "reference" in hl and "preuve" in hl: col_map["ref_preuve"] = i

        cnt = 0
        for row in rows[1:]:
            idx = col_map.get("rs_pro")
            if idx is None or idx >= len(row): continue
            rs = str(row[idx] or "").strip()
            if "eco-volt" not in rs.lower() and "ecovolt" not in rs.lower(): continue
            def gv(k):
                ii = col_map.get(k)
                if ii is not None and ii < len(row):
                    v = row[ii]
                    return str(v).strip() if v is not None else ""
                return ""
            entries.append({
                "sheet": sn, "op_num": gv("op_num"), "rs_pro": rs,
                "siren_pro": gv("siren_pro"), "rs_benef": gv("rs_benef"),
                "ref_interne": gv("ref_interne"), "adresse": gv("adresse"),
                "code_postal": gv("code_postal"), "ville": gv("ville"),
                "telephone": gv("telephone"), "email": gv("email"),
                "ref_preuve": gv("ref_preuve"),
            })
            cnt += 1
        sheet_counts[sn] = cnt
    wb.close()
    return entries, sheet_counts

def fetch_monday():
    items = []
    q = '{ boards(ids: %d) { items_page(limit: 500) { cursor items { id name column_values { id text value } } } } }' % BOARD_ID
    r = requests.post(MONDAY_URL, headers=HEADERS, json={"query": q}, timeout=30).json()
    pg = r["data"]["boards"][0]["items_page"]
    items.extend(pg["items"])
    cursor = pg["cursor"]
    while cursor:
        q2 = '{ next_items_page(limit: 500, cursor: "%s") { cursor items { id name column_values { id text value } } } }' % cursor
        r2 = requests.post(MONDAY_URL, headers=HEADERS, json={"query": q2}, timeout=30).json()
        pg2 = r2["data"]["next_items_page"]
        items.extend(pg2["items"])
        cursor = pg2["cursor"]
        time.sleep(0.3)
    result = []
    for it in items:
        d = {"id": it["id"], "name": it["name"]}
        for cv in it["column_values"]:
            if cv["id"] in MONDAY_COL_MAP:
                d[MONDAY_COL_MAP[cv["id"]]] = (cv.get("text") or "").strip()
        result.append(d)
    return result

def main():
    print("Excel...", file=sys.stderr)
    excel, sc = extract_ecovolt()
    print(f"  {len(excel)} entries {sc}", file=sys.stderr)

    print("Monday...", file=sys.stderr)
    monday = fetch_monday()
    print(f"  {len(monday)} items", file=sys.stderr)

    # Build Monday lookup
    mon_by_name = {}
    for m in monday:
        k = normalize(m["name"])
        mon_by_name.setdefault(k, []).append(m)

    matched = []
    missing = []
    used_ids = set()

    for ex in excel:
        bk = normalize(ex["rs_benef"])
        found = None
        if bk and bk in mon_by_name:
            found = mon_by_name[bk][0]
        if not found and bk and len(bk) > 3:
            for mk, mvs in mon_by_name.items():
                if bk in mk or mk in bk:
                    found = mvs[0]
                    break
        if found:
            used_ids.add(found["id"])
            matched.append((ex, found))
        else:
            missing.append(ex)

    orphans = [m for m in monday if m["id"] not in used_ids]

    # Field comparison
    COMPARE_FIELDS = [
        ("siren_pro", "siren_pro", normalize, normalize),
        ("email", "email", normalize, normalize),
        ("telephone", "telephone", norm_phone, norm_phone),
        ("adresse", "adresse", normalize, normalize),
        ("code_postal", "code_postal", norm_cp, norm_cp),
        ("ville", "ville", normalize, normalize),
    ]

    mismatches = []
    for ex, mon in matched:
        diffs = {}
        for ex_key, mon_key, ex_norm, mon_norm in COMPARE_FIELDS:
            ev = ex_norm(ex.get(ex_key, ""))
            mv = mon_norm(mon.get(mon_key, ""))
            if ev and mv and ev != mv:
                diffs[ex_key] = {"excel": ex.get(ex_key, ""), "monday": mon.get(mon_key, "")}
        if diffs:
            mismatches.append({
                "name": ex["rs_benef"],
                "monday_id": mon["id"],
                "diffs": diffs
            })

    # OUTPUT
    print(f"\n{'='*60}")
    print("ECOVOLT COMPARISON — FINAL REPORT")
    print(f"{'='*60}")
    print(f"Excel total:         {len(excel)} (ZNI:{sc.get('ZNI',0)}, 19-09:{sc.get('19-09-2025',0)}, 01-10:{sc.get('01-10-2025',0)})")
    print(f"Monday total:        {len(monday)}")
    print(f"Matched:             {len(matched)}")
    print(f"With differences:    {len(mismatches)}")
    print(f"Missing from Monday: {len(missing)}")
    print(f"Orphans on Monday:   {len(orphans)}")

    print(f"\n--- ALL ORPHANS ({len(orphans)}) ---")
    for o in orphans:
        print(f"  [{o['id']}] {o['name']}")

    print(f"\n--- MISMATCHES (first 15) ---")
    for m in mismatches[:15]:
        print(f"  {m['name']} [id:{m['monday_id']}]")
        for fld, vals in m["diffs"].items():
            print(f"    {fld}: Excel=\"{vals['excel']}\" vs Monday=\"{vals['monday']}\"")

    print(f"\n--- MISSING FROM MONDAY (first 15) ---")
    for m in missing[:15]:
        print(f"  [{m['sheet']}] op:{m['op_num']} {m['rs_benef']} ({m['email']})")

    # Stats on mismatch fields
    field_counts = {}
    for m in mismatches:
        for fld in m["diffs"]:
            field_counts[fld] = field_counts.get(fld, 0) + 1
    print(f"\n--- MISMATCH BREAKDOWN BY FIELD ---")
    for fld, cnt in sorted(field_counts.items(), key=lambda x: -x[1]):
        print(f"  {fld}: {cnt} items differ")

    # Save full report
    full = {
        "counts": {"excel": len(excel), "monday": len(monday), "matched": len(matched),
                    "mismatched": len(mismatches), "missing": len(missing), "orphans": len(orphans)},
        "sheet_counts": sc,
        "orphans": [{"id": o["id"], "name": o["name"]} for o in orphans],
        "mismatches_all": mismatches,
        "missing_all": [{"sheet": m["sheet"], "op": m["op_num"], "benef": m["rs_benef"],
                         "email": m["email"], "tel": m["telephone"]} for m in missing],
        "field_mismatch_counts": field_counts,
    }
    out = "/Users/john/JARVIS/projets/velo/scripts/ecovolt-comparison-full.json"
    with open(out, "w") as f:
        json.dump(full, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nFull report: {out}", file=sys.stderr)

if __name__ == "__main__":
    main()
